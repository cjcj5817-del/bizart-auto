# -*- coding: utf-8 -*-
import streamlit as st
import openpyxl
import io
import re
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ── 설정 ──────────────────────────────────────────────────────────
MIN_VM      = 500_000
SUB_MONTHS  = 23
MAIL_TO1    = "cjcj5817@naver.com"   # 테스트용 (실제: artmanse21@naver.com)
MAIL_TO2    = "cjcj5817@naver.com"   # 테스트용 (실제: seouldm@seouldm.co.kr)
SMTP_SERVER = "api.mail.bizbee.co.kr"
SMTP_PORT   = 587

def add_months(yyyymm, m):
    y, mo = yyyymm // 100, yyyymm % 100
    t = y * 12 + (mo - 1) + m
    return (t // 12) * 100 + (t % 12 + 1)

def sv(v):
    return '' if v is None else str(v).strip()

def nv(v):
    try: return float(v) if v is not None else 0
    except: return 0

# ── 핵심 처리 ─────────────────────────────────────────────────────
def process_from_files(ke_bytes, km_bytes, jw_bytes, addr_bytes, existing_bytes, mak):
    """
    ke_bytes   : 계약등록 xlsx (신계약 시트)
    km_bytes   : 계약관리 xlsx (보유계약 시트) - optional
    jw_bytes   : 26.XX월 지원물품 xlsx (키 시트 + 인사정보목록 시트)
    addr_bytes : 업체주소 xlsx (A열=증권번호, D열=주소)
    existing_bytes : 기존 발송명단 xlsx - optional (중복 제외)
    mak        : 마감월 int (YYYYMM)
    """
    log = []
    now = datetime.now()
    log.append(f"실행: {now.strftime('%Y-%m-%d %H:%M:%S')}")
    log.append(f"마감월: {mak}  →  최종발송호: {add_months(mak, SUB_MONTHS)}")

    # ── 1. 키 시트 로드 (지원물품) ──────────────────────────────
    key_map = {}
    hr_map  = {}
    if jw_bytes:
        wb_jw = openpyxl.load_workbook(io.BytesIO(jw_bytes), data_only=True)
        for row in wb_jw['키'].iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                key_map[sv(row[0])] = int(row[1]) if row[1] is not None else 0
        log.append(f"키 시트: {len(key_map)}개 계약상태")

        # ── 2. 인사정보목록 로드 (지원물품) → 사번: (PM명, 주소) ──
        for row in wb_jw['인사정보목록'].iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                사번 = sv(row[0])
                pm명 = sv(row[1])
                주소 = sv(row[38]) if len(row) > 38 else ''
                hr_map[사번] = (pm명, 주소)
        log.append(f"인사정보목록: {len(hr_map)}명")
    else:
        log.append("지원물품: 파일 없음 (계약상태 키/인사정보 미반영)")

    # ── 3. 업체주소 로드 → 증권번호: 주소 ──────────────────────
    addr_map = {}
    if addr_bytes:
        wb_addr  = openpyxl.load_workbook(io.BytesIO(addr_bytes), data_only=True)
        ws_addr  = wb_addr.active
        for row in ws_addr.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                증권 = sv(row[0])
                주소 = sv(row[3]) if len(row) > 3 else ''
                if 증권 and 주소:
                    addr_map[증권] = 주소
        log.append(f"업체주소: {len(addr_map)}개 업체")
    else:
        log.append("업체주소: 파일 없음 (받는주소 모두 공란 처리)")

    # ── 4. 계약관리 로드 → 증권번호: 최신 계약상태 (선택) ──────
    status_map = {}
    if km_bytes:
        wb_km = openpyxl.load_workbook(io.BytesIO(km_bytes), data_only=True)
        ws_km = wb_km['보유계약']
        for row in ws_km.iter_rows(min_row=4, values_only=True):
            if row[7] is not None:   # 증권번호 = col 8 (index 7)
                증권  = sv(row[7])
                상태  = sv(row[22])  # 계약상태 = col 23 (index 22)
                if 증권: status_map[증권] = 상태
        log.append(f"계약관리(보유계약): {len(status_map)}건 계약상태 로드")

    # ── 5. 기존 발송명단 로드 → 전체 행 보존 + 증권번호 중복 체크 ──
    existing_rows = []   # 기존 데이터 전체 (헤더 제외)
    existing_set  = set()
    if existing_bytes:
        wb_ex = openpyxl.load_workbook(io.BytesIO(existing_bytes), data_only=True)
        ws_ex = wb_ex.active
        for row in ws_ex.iter_rows(min_row=2, values_only=True):
            증권v = sv(row[9]) if len(row) > 9 else ''
            if not any(c is not None for c in row): continue  # 빈 행 스킵
            existing_rows.append(list(row))
            if 증권v: existing_set.add(증권v)
        log.append(f"기존 발송명단: {len(existing_rows)}건 (그대로 포함)")

    # ── 6. 계약등록 신계약 처리 ─────────────────────────────────
    wb_ke = openpyxl.load_workbook(io.BytesIO(ke_bytes), data_only=True)
    ws_ke = wb_ke['신계약']

    total = ok = ex_cont = ex_amt = no_addr_cnt = na_send = dup = 0
    new_entries = []

    for row in ws_ke.iter_rows(min_row=4, values_only=True):
        증권 = sv(row[6])   # 증권번호 col 7
        if not 증권: continue

        모집     = sv(row[0])   # 모집자명 col 1
        사번     = sv(row[1])   # 모집자사번 col 2
        계약자   = sv(row[7])   # 계약자 col 8
        vm       = nv(row[12])  # VM환산 col 13 (M열)

        # 계약상태: 계약관리에 있으면 최신값 사용, 없으면 계약등록 값
        계약상태 = status_map.get(증권, sv(row[14]))  # col 15

        # key_map이 비어있으면(지원물품 미업로드) 계약상태 필터 미적용
        유지     = '대상' if (not key_map or key_map.get(계약상태, 0) == 1) else '제외'
        금액     = '대상' if vm >= MIN_VM else '제외'
        받는주소 = addr_map.get(증권, '')
        최종     = '대상' if (유지 == '대상' and 금액 == '대상') else '제외'

        hr           = hr_map.get(사번, ('', '#N/A'))
        pm명         = hr[0] if hr[0] else 모집
        보내는주소   = hr[1] if hr[1] else '#N/A'
        업체명       = f"{계약자} 대표님" if 계약자 else ''

        total += 1
        if 최종 == '대상':
            if not 받는주소.strip():
                no_addr_cnt += 1
            elif 보내는주소 == '#N/A':
                na_send += 1
            elif 증권 in existing_set:
                dup += 1  # 기존 명단에 이미 있음
            else:
                ok += 1
                new_entries.append({
                    '사번': 사번, '모집': pm명,
                    '보내는주소': 보내는주소, '업체명': 업체명,
                    '받는주소': 받는주소, '증권': 증권,
                    '최초': mak, '최종': add_months(mak, SUB_MONTHS)
                })
        else:
            if 유지 == '제외': ex_cont += 1
            if 금액 == '제외': ex_amt  += 1

    log.append(f"계약등록 처리 — 전체: {total}건 | 신규추가: {ok}건 | "
                f"계약상태제외: {ex_cont}건 | 금액미달: {ex_amt}건 | "
                f"주소없음: {no_addr_cnt}건 | 인사없음: {na_send}건 | 중복(기존포함): {dup}건")

    total_out = len(existing_rows) + ok
    log.append(f"최종 발송명단 — 기존 {len(existing_rows)}건 + 신규 {ok}건 = 총 {total_out}건")

    # ── 7. 발주 첨부 xlsx 생성 (기존 + 신규 합산) ────────────────
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "발송명단"
    headers = ['최초발송호', '최종발송호', '사번', '보내는사람', '보내는주소',
               '업체명', '받는주소', '분류', '중기이코노미기업지원단', '증권번호']
    for ci, h in enumerate(headers, 1):
        ws_out.cell(1, ci).value = h

    # 기존 발송명단 행 그대로 복사
    ri = 2
    for row in existing_rows:
        for ci, val in enumerate(row[:10], 1):
            ws_out.cell(ri, ci).value = val
        ri += 1

    # 신규 추가분 append
    for e in new_entries:
        ws_out.cell(ri,  1).value = e['최초']
        ws_out.cell(ri,  2).value = e['최종']
        ws_out.cell(ri,  3).value = e['사번']
        ws_out.cell(ri,  4).value = e['모집']
        ws_out.cell(ri,  5).value = e['보내는주소']
        ws_out.cell(ri,  6).value = e['업체명']
        ws_out.cell(ri,  7).value = e['받는주소']
        ws_out.cell(ri,  8).value = '중기이코노미기업지원단'
        ws_out.cell(ri,  9).value = 1
        ws_out.cell(ri, 10).value = e['증권']
        ri += 1

    buf_out = io.BytesIO()
    wb_out.save(buf_out)
    buf_out.seek(0)

    # 파일명 생성
    yy = str(now.year)[2:]
    mm = f"{now.month:02d}"
    mak_yy = str(mak // 100)[2:]
    mak_mm = f"{mak % 100:02d}"
    attach_name = f"{yy}.{mm}월호 비자트 ({mak_yy}.{mak_mm}마감).xlsx"

    # 메일 내용
    send_date    = now.strftime('%Y.%m.%d')
    mail_subject = f"(주)밸류마크 {yy}.{mm}월호 비자트 발주 내용 전달 ({send_date})"
    mail_body    = f"안녕하세요. 밸류마크 총무팀입니다.\n{yy}년 {mm}월호 비자트 발주 내용 전달드립니다.\n감사합니다."

    log.append(f"발주 엑셀 생성 — {attach_name} (총 {total_out}건)")

    stats = dict(
        total=total, ok=ok, total_out=total_out,
        existing=len(existing_rows),
        ex_cont=ex_cont, ex_amt=ex_amt,
        no_addr=no_addr_cnt, na_send=na_send, dup=dup,
        attach_name=attach_name,
        mail_subject=mail_subject, mail_body=mail_body,
        log=log
    )
    return buf_out.read(), stats

# ── SMTP 발송 ──────────────────────────────────────────────────────
def send_mail(mail_from, mail_pw, subject, body, attach_bytes, attach_name):
    msg = MIMEMultipart()
    msg['From']    = mail_from
    msg['To']      = f"{MAIL_TO1}, {MAIL_TO2}"
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain', 'utf-8'))
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attach_bytes)
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', attach_name))
    msg.attach(part)
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=15) as server:
        server.ehlo(); server.starttls(); server.ehlo()
        server.login(mail_from, mail_pw)
        server.sendmail(mail_from, [MAIL_TO1, MAIL_TO2], msg.as_bytes())

# ══════════════════════════════════════════════════════════════════
# UI
# ══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="비자트 자동화", page_icon="📮", layout="wide")

st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f8fafc; }
[data-testid="stSidebar"] { background: #1e3a5f; }
[data-testid="stSidebar"] * { color: white !important; }
.upload-box {
    background: white; border-radius: 12px; padding: 20px;
    border: 1px solid #e2e8f0; margin-bottom: 12px;
}
</style>
""", unsafe_allow_html=True)

# ── 로그인 ────────────────────────────────────────────────────────
if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    col_l, col_m, col_r = st.columns([1, 1.2, 1])
    with col_m:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("## 📮 비자트 자동화")
        st.markdown("**(주)밸류마크 총무팀**")
        pw = st.text_input("비밀번호", type="password", label_visibility="collapsed",
                           placeholder="접속 비밀번호 입력")
        if st.button("로그인", use_container_width=True, type="primary"):
            if pw == st.secrets.get("app_password", "valuemark2026"):
                st.session_state.auth = True
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    st.stop()

# ── 사이드바 ──────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📮 비자트 자동화")
    st.markdown("**(주)밸류마크 총무팀**")
    st.divider()
    st.markdown("### 📋 매달 업로드 파일")
    st.markdown("""
① 계약등록 xlsx *(필수)*\n
② 계약관리 xlsx\n
③ 지원물품 xlsx\n
④ 업체주소 xlsx *(유창식 이사님)*\n
⑤ 기존 발송명단 xlsx
    """)
    st.divider()
    st.markdown("### ✅ 자동 처리")
    st.markdown("""
- 계약상태 유지 여부 확인
- VM환산 50만원 이상 확인
- 업체주소 매핑
- 인사정보(PM 주소) 매핑
- 중복 계약 제외
- 발주 엑셀 생성
- 발주 메일 초안
    """)
    st.divider()
    if st.button("🔒 로그아웃"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()

# ── 메인 ─────────────────────────────────────────────────────────
st.title("📮 비자트 발송 자동화")
st.caption(f"실행일: {datetime.now().strftime('%Y년 %m월 %d일')}  |  (주)밸류마크 총무팀")
st.divider()

# ── 파일 업로드 ────────────────────────────────────────────────
st.markdown("### 📁 파일 업로드")
st.caption("아래 파일들을 업로드하고 자동화 실행 버튼을 클릭하세요.")

col1, col2 = st.columns(2)
with col1:
    f_ke   = st.file_uploader("① 계약등록 xlsx",   type=["xlsx"], key="ke",
                               help="신계약 시트가 포함된 계약등록 파일")
    f_jw   = st.file_uploader("③ 지원물품 xlsx",   type=["xlsx"], key="jw",
                               help="키 시트 + 인사정보목록 시트 포함")
    f_exist= st.file_uploader("⑤ 기존 발송명단 xlsx *(선택)*", type=["xlsx"], key="exist",
                               help="이전 달까지의 발송명단 - 중복 제외에 사용")
with col2:
    f_km   = st.file_uploader("② 계약관리 xlsx *(선택)*",   type=["xlsx"], key="km",
                               help="보유계약 시트 - 최신 계약상태 반영")
    f_addr = st.file_uploader("④ 업체주소 xlsx *(선택)*",   type=["xlsx"], key="addr",
                               help="유창식 이사님께 받은 업체주소 파일 (A열=증권번호, D열=주소)")

st.divider()

# ── 설정 ──────────────────────────────────────────────────────────
st.markdown("### ⚙️ 설정")
col_s1, col_s2, col_s3 = st.columns(3)
with col_s1:
    now = datetime.now()
    default_mak = f"{now.year}{now.month:02d}"
    mak_input = st.text_input("📅 마감월 (YYYYMM)", value=default_mak,
                               help="예: 2026년 3월 마감 → 202603")
with col_s2:
    mail_from = st.text_input("📧 사내 이메일 주소", placeholder="hong@valuemark.co.kr")
with col_s3:
    mail_pw   = st.text_input("🔑 사내 메일 비밀번호", type="password",
                               help="입력하면 완료 후 자동 발송됩니다.")

# ── 실행 ──────────────────────────────────────────────────────────
if not f_ke:
    st.info("① 계약등록 파일을 업로드하면 실행 버튼이 활성화됩니다.")
else:
    if st.button("🚀 자동화 실행", type="primary", use_container_width=True):
        try:
            mak = int(mak_input)
            assert 200001 <= mak <= 209912
        except:
            st.error("마감월 형식이 올바르지 않습니다. 예: 202603")
            st.stop()

        with st.spinner("⏳ 처리 중입니다. 잠시만 기다려주세요..."):
            xlsx_out, stats = process_from_files(
                ke_bytes      = f_ke.read(),
                km_bytes      = f_km.read()     if f_km    else None,
                jw_bytes      = f_jw.read()     if f_jw    else None,
                addr_bytes    = f_addr.read()   if f_addr  else None,
                existing_bytes= f_exist.read()  if f_exist else None,
                mak           = mak
            )

        st.session_state.update({
            "xlsx_out":    xlsx_out,
            "stats":       stats,
            "mail_from":   mail_from,
            "mail_pw":     mail_pw,
            "processed":   True,
            "mail_sent":   False,
            "mail_error":  "",
        })

        # 이메일 + 비밀번호 있으면 즉시 발송
        if mail_from and mail_pw:
            with st.spinner("📨 메일 발송 중..."):
                try:
                    send_mail(mail_from, mail_pw, stats["mail_subject"],
                              stats["mail_body"], xlsx_out, stats["attach_name"])
                    st.session_state["mail_sent"] = True
                except smtplib.SMTPAuthenticationError:
                    st.session_state["mail_error"] = "로그인 실패 — 이메일/비밀번호를 확인해주세요."
                except Exception as e:
                    st.session_state["mail_error"] = str(e)

# ── 결과 ──────────────────────────────────────────────────────────
if st.session_state.get("processed"):
    stats     = st.session_state["stats"]
    xlsx_out  = st.session_state["xlsx_out"]
    mail_sent = st.session_state.get("mail_sent", False)
    mail_err  = st.session_state.get("mail_error", "")

    st.divider()

    if mail_sent:
        st.success(f"✅ 자동화 완료!  |  📨 메일 발송 완료 ({MAIL_TO1}, {MAIL_TO2})")
    else:
        st.success("✅ 자동화 완료!")

    if mail_err:
        st.error(f"❌ 메일 발송 실패: {mail_err}")

    # 핵심 지표
    st.markdown("### 📊 처리 결과")
    c0, c1, c2 = st.columns(3)
    c0.metric("📋 최종 발송명단 합계", f"{stats['total_out']}건",
              help="기존 + 신규 합산")
    c1.metric("기존 발송명단",  f"{stats['existing']}건")
    c2.metric("신규 추가",      f"{stats['ok']}건")

    st.markdown("##### 신규 계약 검토 결과")
    d1, d2, d3, d4, d5 = st.columns(5)
    d1.metric("신규 계약 전체",  f"{stats['total']}건")
    d2.metric("계약상태 제외",   f"{stats['ex_cont']}건")
    d3.metric("금액 미달 제외",  f"{stats['ex_amt']}건")
    d4.metric("주소 없어 제외",  f"{stats['no_addr']}건")
    d5.metric("중복(기존포함)",  f"{stats['dup']}건")

    if stats['na_send'] > 0:
        st.warning(f"⚠️ 인사정보 없음 {stats['na_send']}건 — 인사정보목록 최신 업데이트 필요")

    st.divider()

    # 다운로드
    st.markdown("### 📥 파일 다운로드")
    st.download_button(
        label=f"📎 발주 첨부 xlsx 다운로드 — {stats['attach_name']}",
        data=xlsx_out,
        file_name=stats["attach_name"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.divider()

    # 발주 메일
    st.markdown("### 📧 발주 메일")
    col_m1, col_m2 = st.columns(2)
    with col_m1:
        st.markdown("**수신인**")
        st.markdown(f"- 예술만세: `{MAIL_TO1}`")
        st.markdown(f"- 서울DM: `{MAIL_TO2}`")
    with col_m2:
        st.markdown("**첨부파일**")
        st.markdown(f"📎 `{stats['attach_name']}`")

    st.markdown("**제목** *(우측 아이콘으로 복사)*")
    st.code(stats["mail_subject"], language=None)
    st.markdown("**본문** *(우측 아이콘으로 복사)*")
    st.code(stats["mail_body"], language=None)

    if not mail_sent:
        st.markdown("---")
        st.markdown("**📨 메일 직접 발송**")
        col_a, col_b = st.columns(2)
        with col_a:
            from2 = st.text_input("사내 이메일 주소", placeholder="hong@valuemark.co.kr", key="from2")
        with col_b:
            pw2   = st.text_input("사내 메일 비밀번호", type="password", key="pw2")
        if st.button("📨 메일 발송", type="primary"):
            if from2 and pw2:
                with st.spinner("발송 중..."):
                    try:
                        send_mail(from2, pw2, stats["mail_subject"],
                                  stats["mail_body"], xlsx_out, stats["attach_name"])
                        st.success(f"✅ 발송 완료! 수신: {MAIL_TO1}, {MAIL_TO2}")
                    except smtplib.SMTPAuthenticationError:
                        st.error("❌ 로그인 실패 — 이메일/비밀번호를 확인해주세요.")
                    except Exception as e:
                        st.error(f"❌ 발송 실패: {e}")
            else:
                st.warning("이메일 주소와 비밀번호를 입력해주세요.")

    # 처리 로그
    with st.expander("📋 상세 처리 로그", expanded=(stats['total_out'] == 0)):
        for line in stats["log"]:
            st.text(line)
    if stats['total_out'] == 0:
        st.warning("⚠️ 최종 발송명단이 0건입니다. 위 로그에서 원인을 확인해주세요.")
